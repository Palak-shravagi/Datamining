import openpyxl
import math
wb = openpyxl.load_workbook("input.xlsx")
sheet_obj = wb.active
row = sheet_obj.max_row
col = sheet_obj.max_column


def probO(arr, l):
    prob = 0
    for i in range(0, l):
        if arr[i] == 1:
            prob = prob+1
    return prob


def prob_A_B(arr, arr1, l):
    prob = 0
    for i in range(0, l):
        if arr[i] == arr1[i] == 1:
            prob = prob+1
    return prob


r1 = int(input("Enter first row no:"))
r2 = int(input("Enter second row no:"))
A = []
B = []
for i in range(2, col+1):
    val1 = sheet_obj.cell(row=r1, column=i)
    val2 = sheet_obj.cell(row=r2, column=i)
    A.append(val1.value)
    B.append(val2.value)

corr = prob_A_B(A, B, col-1)/(probO(A, col-1)*probO(B, col-1))
ansVal = sheet_obj.cell(row=row+2, column=2)
ansVal.value = corr
tmp = sheet_obj.cell(row=row+2, column=1)
tmp.value = "Correlation of row:"+str(r1)+" and row:"+str(r2)+" "
wb.save("output.xlsx")
