
import openpyxl
import math


def cal_median(array, l, r):
    n = r+l
    median = 0
    if n % 2 == 0:
        median = array[n//2]
    else:
        median1 = array[n//2]
        median2 = array[(n//2) + 1]

        median = (median1 + median2)/2
    return median


def five_no_summary(Arr, Ans):
    min = Arr[0]
    max = Arr[0]
    for x in Arr:
        if (x < min):
            min = x
    for x in Arr:
        if (x > max):
            max = x

    Arr.sort()
    n = len(Arr)
    # Median
    Med = cal_median(Arr, 0, n-1)
    # Q1,Q3
    Q1 = 0
    Q3 = 0
    if (n % 2 == 0):
        Q1 = cal_median(Arr, 0, (n//2)-1)
        Q3 = cal_median(Arr, n//2, n-1)
    else:
        Q1 = cal_median(Arr, 0, (n//2)-1)
        Q3 = cal_median(Arr, (n+1)//2, n-1)
    Ans.append(min)
    Ans.append(Q1)
    Ans.append(Med)
    Ans.append(Q3)
    Ans.append(max)


wb = openpyxl.load_workbook(
    "numsum_.xlsx")
sheet_obj = wb.active
row = sheet_obj.max_row
col = sheet_obj.max_column

i = 1

# for i in range(1,col+1):
arr = []
ans = []
for j in range(2, row+1):
    val = sheet_obj.cell(row=j, column=i)
    arr.append(val.value)

five_no_summary(arr, ans)
i = 0
j = 2

arr1 = ["Min", "Quatrile1", "Median/Quatrile2", "Quatrile3", "Max"]

for i in range(0, 5):
    c0 = sheet_obj.cell(row=j, column=3)
    c1 = sheet_obj.cell(row=j, column=4)

    c0.value = arr1[i]
    c1.value = ans[i]
    i = i+1
    j = j+1

wb.save("numsum_.xlsx")
